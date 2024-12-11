import customtkinter as ctk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
import pywhatkit
from datetime import datetime


class MetroBookingApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Metro Ticket Booking")
        self.geometry("700x600")

        # Create Excel file if not exists
        self.file_name = "MetroBookings.xlsx"
        if not os.path.exists(self.file_name):
            self.create_excel_file()

        # Title Label
        self.title_label = ctk.CTkLabel(self, text="Metro Ticket Booking", font=("Arial", 24, "bold"))
        self.title_label.pack(pady=20)

        # Station Selection
        self.source_label = ctk.CTkLabel(self, text="Source Station:", font=("Arial", 14))
        self.source_label.pack(pady=5)
        self.source_dropdown = ctk.CTkOptionMenu(
            self, values=["Station A", "Station B", "Station C", "Station D", "Station E"]
        )
        self.source_dropdown.pack(pady=5)

        self.destination_label = ctk.CTkLabel(self, text="Destination Station:", font=("Arial", 14))
        self.destination_label.pack(pady=5)
        self.destination_dropdown = ctk.CTkOptionMenu(
            self, values=["Station A", "Station B", "Station C", "Station D", "Station E"]
        )
        self.destination_dropdown.pack(pady=5)

        # Number of Tickets
        self.tickets_label = ctk.CTkLabel(self, text="Number of Tickets:", font=("Arial", 14))
        self.tickets_label.pack(pady=5)
        self.tickets_spinbox = ctk.CTkEntry(self, width=300, placeholder_text="Enter number of tickets")
        self.tickets_spinbox.pack(pady=5)

        # WhatsApp Number
        self.whatsapp_label = ctk.CTkLabel(self, text="WhatsApp Number (with country code):", font=("Arial", 14))
        self.whatsapp_label.pack(pady=5)
        self.whatsapp_entry = ctk.CTkEntry(self, width=300, placeholder_text="+91XXXXXXXXXX")
        self.whatsapp_entry.pack(pady=5)

        # Button for Booking
        self.book_button = ctk.CTkButton(self, text="Book Tickets", command=self.book_ticket, width=200)
        self.book_button.pack(pady=20)

        # Reset Button
        self.reset_button = ctk.CTkButton(self, text="Reset", command=self.reset_fields, width=200, fg_color="red")
        self.reset_button.pack(pady=10)

        # Footer
        self.footer_label = ctk.CTkLabel(
            self,
            text="© 2024 Metro Booking App | All Rights Reserved",
            font=("Arial", 12),
        )
        self.footer_label.pack(side="bottom", pady=10)

    def create_excel_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"
        ws.append(["Source", "Destination", "Tickets", "WhatsApp Number", "Date/Time"])
        wb.save(self.file_name)

    def save_to_excel(self, source, destination, tickets, whatsapp_number):
        wb = load_workbook(self.file_name)
        ws = wb.active
        ws.append([source, destination, tickets, whatsapp_number, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(self.file_name)

    def send_whatsapp_message(self, number, source, destination, tickets):
        try:
            message = (
                f"Metro Ticket Booking Confirmed!\n\n"
                f"Source: {source}\n"
                f"Destination: {destination}\n"
                f"Tickets: {tickets}\n\n"
                f"Thank you for booking with us!"
            )
            # Schedule a message 1 minute from now
            pywhatkit.sendwhatmsg_instantly(number, message, wait_time=20, tab_close=True)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to send WhatsApp message: {e}")

    def book_ticket(self):
        source = self.source_dropdown.get()
        destination = self.destination_dropdown.get()
        tickets = self.tickets_spinbox.get()
        whatsapp_number = self.whatsapp_entry.get()

        if not tickets or not whatsapp_number:
            messagebox.showerror("Error", "All fields are required!")
            return

        if not tickets.isdigit() or int(tickets) <= 0:
            messagebox.showerror("Error", "Enter a valid number of tickets!")
            return

        if source == destination:
            messagebox.showerror("Error", "Source and destination cannot be the same!")
            return

        if not whatsapp_number.startswith("+") or not whatsapp_number[1:].isdigit():
            messagebox.showerror("Error", "Enter a valid WhatsApp number with country code!")
            return

        # Save booking to Excel
        self.save_to_excel(source, destination, tickets, whatsapp_number)

        # Send WhatsApp message
        self.send_whatsapp_message(whatsapp_number, source, destination, tickets)

        messagebox.showinfo(
            "Booking Confirmed",
            f"Tickets booked successfully!\n\nSource: {source}\nDestination: {destination}\nTickets: {tickets}",
        )
        self.reset_fields()

    def reset_fields(self):
        self.source_dropdown.set("Station A")
        self.destination_dropdown.set("Station A")
        self.tickets_spinbox.delete(0, "end")
        self.whatsapp_entry.delete(0, "end")


if __name__ == "__main__":
    app = MetroBookingApp()
    app.mainloop()
